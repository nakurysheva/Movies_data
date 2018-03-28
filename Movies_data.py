from bs4 import BeautifulSoup
from urllib.request import urlopen
from openpyxl import Workbook

def namecells():# The fuction to name cells in the Excel sheet
    ws['A1'] = "Movie Name"
    ws['B1'] = "Poster link"
    ws['C1'] = "Logline"
    ws['D1'] = "Rating"


def stringOnly(htmlstr, catname): #The function picks strings from a piece of html tree and returns this string under necessary category name
    for i in htmlstr:
        catname = (str(i))
    return catname


def linkOnly(htmltree):  #This function extracts link from html tag
    for link in htmltree('a'):
        return 'http://www.imdb.com' + str(link.get('href'))


def movieInfoInImdb(data):  #The function to parse necessary info from html tree
    moviename_tag = data.a
    Name = ''
    Name = stringOnly(moviename_tag, Name)
    rating_block = data.find('div', 'ratings-bar')
    rating_tag = rating_block.strong
    Rating = ''
    Rating = stringOnly(rating_tag, Rating)
    summary_info = data.find_all('p', class_="text-muted")
    summary_info = summary_info[1]
    Summary = ''
    Summary = stringOnly(summary_info, Summary)
    Summary = Summary.strip()
    return Name, Summary, Rating



wb = Workbook()
ws = wb.get_sheet_by_name("Sheet")
namecells()

data = urlopen("http://www.imdb.com/search/title?genres=film_noir").read()
soup = BeautifulSoup(data, "html5lib")
name_block = soup.find_all('div', class_="lister-item-content")

# The folowing code extracts links to the movie page from the html tree
poster_block = soup.find('div', class_="lister-item-image float-left")
link = (linkOnly(poster_block))
count = 2
ws['B' + str(count)] = link
poster_count += 1
while poster_count <= 51:
    poster_block = poster_block.find_next('div', class_="lister-item-image float-left")
    link = (linkOnly(poster_block))
    if link != None:
        ws['B' + str(poster_count)] = link
    else:
        ws['B' + str(poster_count)] = None
    poster_count += 1

#The following code starts the movieInfoInImdb function to fill in necessary cells
count = 2
for i in name_block:
    ws['A' + str(count)], ws['C' + str(count)], ws['D' + str(count)] = movieInfoInImdb(i)
    count += 1
wb.save('Movies data.xlsx')
